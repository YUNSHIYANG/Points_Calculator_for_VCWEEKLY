"""
Excel读写模块

提供Excel文件的读写功能。
"""

import openpyxl
from pathlib import Path
from typing import Dict, Optional, Tuple
import logging

try:
    from .models import VideoStats
except ImportError:
    from weekly_score.core.models import VideoStats

logger = logging.getLogger(__name__)


class ExcelError(Exception):
    """Excel操作异常基类"""
    pass


class ExcelReadError(ExcelError):
    """Excel读取异常"""
    pass


class ExcelWriteError(ExcelError):
    """Excel写入异常"""
    pass


class ExcelManager:
    """Excel文件管理器"""
    
    def __init__(self, file_path: Path, sheet_name: str = "Sheet1"):
        """
        初始化Excel管理器
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
        """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self._workbook = None
        self._worksheet = None
    
    def open(self, data_only: bool = False):
        """
        打开Excel文件
        
        Args:
            data_only: 是否以只读数据模式打开（读取公式计算值）
        """
        try:
            if not self.file_path.exists():
                raise ExcelReadError(f"文件不存在: {self.file_path}")
            
            self._workbook = openpyxl.load_workbook(self.file_path, data_only=data_only)
            self._worksheet = self._workbook[self.sheet_name]
            logger.debug(f"已打开Excel文件: {self.file_path}")
        except KeyError:
            raise ExcelReadError(f"工作表不存在: {self.sheet_name}")
        except Exception as e:
            raise ExcelReadError(f"打开文件失败: {e}")
    
    def close(self):
        """关闭Excel文件"""
        if self._workbook:
            self._workbook.close()
            self._workbook = None
            self._worksheet = None
            logger.debug(f"已关闭Excel文件: {self.file_path}")
    
    def __enter__(self):
        self.open()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
    
    def read_base_stats(self, start_row: int = 11) -> VideoStats:
        """
        读取基数数据（历史累计）
        
        从第 start_row 行开始遍历，逐行累加直到遇到空行。
        此逻辑替代读取 SUM 公式，避免公式缓存问题。
        
        Args:
            start_row: 数据起始行号，默认为11
            
        Returns:
            VideoStats: 历史累计基数
            
        Raises:
            ExcelReadError: 读取失败
        """
        if not self._worksheet:
            raise ExcelReadError("Excel文件未打开")
        
        try:
            total_view = 0
            total_like = 0
            total_danmaku = 0
            total_reply = 0
            total_coin = 0
            total_favorite = 0
            
            row = start_row
            while True:
                # 读取B列（播放量）判断是否为空行
                view_val = self._worksheet.cell(row, 2).value
                if view_val is None or (isinstance(view_val, str) and view_val.strip() == ''):
                    break
                
                total_view += self._safe_float(view_val)
                total_like += self._safe_float(self._worksheet.cell(row, 3).value)
                total_danmaku += self._safe_float(self._worksheet.cell(row, 4).value)
                total_reply += self._safe_float(self._worksheet.cell(row, 5).value)
                total_coin += self._safe_float(self._worksheet.cell(row, 6).value)
                total_favorite += self._safe_float(self._worksheet.cell(row, 7).value)
                
                row += 1
            
            base = VideoStats(
                view=total_view,
                like=total_like,
                danmaku=total_danmaku,
                reply=total_reply,
                coin=total_coin,
                favorite=total_favorite
            )
            logger.debug(f"已读取基数数据（从第{start_row}行累加）: {base}")
            return base
        except Exception as e:
            raise ExcelReadError(f"读取基数数据失败: {e}")
    
    def write_stats(self, stats: VideoStats, beijing_time_str: str, start_row: int = 11) -> bool:
        """
        写入统计数据（追加到下一个空行）
        
        Args:
            stats: 视频统计数据
            beijing_time_str: 北京时间字符串
            start_row: 数据起始行号，默认为11
            
        Returns:
            bool: 是否写入成功
            
        Raises:
            ExcelWriteError: 写入失败
        """
        if not self._worksheet:
            raise ExcelReadError("Excel文件未打开")
        
        try:
            # 写入时间信息到合并单元格A1:G1
            self._write_time_info(beijing_time_str)
            
            # 查找下一个空行（从 start_row 开始）
            target_row = self._find_next_empty_row(start_row)
            
            # 写入数据
            self._worksheet.cell(target_row, 2, stats.view)
            self._worksheet.cell(target_row, 3, stats.like)
            self._worksheet.cell(target_row, 4, stats.danmaku)
            self._worksheet.cell(target_row, 5, stats.reply)
            self._worksheet.cell(target_row, 6, stats.coin)
            self._worksheet.cell(target_row, 7, stats.favorite)
            
            # 保存文件
            self._workbook.save(self.file_path)
            logger.info(f"已写入统计数据到第{target_row}行: {self.file_path}")
            return True
            
        except ExcelWriteError:
            raise
        except Exception as e:
            raise ExcelWriteError(f"写入数据失败: {e}")
    
    def _write_time_info(self, beijing_time_str: str):
        """
        写入时间信息
        
        Args:
            beijing_time_str: 北京时间字符串
        """
        # 检查A1:G1是否已合并
        if 'A1:G1' not in [str(mr) for mr in self._worksheet.merged_cells]:
            self._worksheet.merge_cells('A1:G1')
        
        self._worksheet.cell(1, 1).value = f"数据获取时间：{beijing_time_str}"
    
    def _get_cell_value(self, row: int, col: int, default: int = 0) -> int:
        """
        获取单元格值
        
        Args:
            row: 行号
            col: 列号
            default: 默认值
            
        Returns:
            单元格值
        """
        value = self._worksheet.cell(row, col).value
        if value is None:
            return default
        
        # 确保返回整数类型
        try:
            return int(value)
        except (ValueError, TypeError):
            logger.warning(f"单元格 ({row}, {col}) 的值 '{value}' 无法转换为整数，使用默认值")
            return default
    
    def _safe_float(self, value) -> float:
        """
        安全转换为浮点数
        
        Args:
            value: 待转换的值
            
        Returns:
            转换后的浮点数
        """
        if value is None:
            return 0.0
        if isinstance(value, str):
            value = value.strip()
            if value == '':
                return 0.0
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0
    
    def _find_next_empty_row(self, start_row: int = 11) -> int:
        """
        从指定行开始查找下一个空行
        
        Args:
            start_row: 起始行号
            
        Returns:
            下一个空行的行号
        """
        row = start_row
        while True:
            value = self._worksheet.cell(row, 2).value
            if value is None or (isinstance(value, str) and value.strip() == ''):
                return row
            row += 1
    
    def _is_cell_merged(self, row: int, col: int) -> bool:
        """
        检查单元格是否被合并
        
        Args:
            row: 行号
            col: 列号
            
        Returns:
            是否被合并
        """
        for merged_range in self._worksheet.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and 
                merged_range.min_col <= col <= merged_range.max_col):
                return True
        return False