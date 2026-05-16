"""
Excel模块测试

测试Excel读写功能。
"""

import pytest
import tempfile
import os
from pathlib import Path
from unittest.mock import patch, MagicMock
import openpyxl

from weekly_score.core.excel_io import ExcelManager, ExcelError, ExcelReadError, ExcelWriteError
from weekly_score.core.models import VideoStats


class TestExcelManager:
    """ExcelManager测试类"""
    
    def setup_method(self):
        """测试前设置"""
        # 创建临时目录
        self.temp_dir = tempfile.mkdtemp()
        self.test_file = Path(self.temp_dir) / "test.xlsx"
        
        # 创建测试Excel文件
        self.create_test_excel()
    
    def teardown_method(self):
        """测试后清理"""
        # 清理临时文件
        if self.test_file.exists():
            os.remove(self.test_file)
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)
    
    def create_test_excel(self):
        """创建测试Excel文件"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # 写入测试数据：模拟周刊格式
        # 第11行起为历史数据
        ws.cell(11, 2, 50000)   # B11: 播放量
        ws.cell(11, 3, 2000)    # C11: 点赞量
        ws.cell(11, 4, 400)     # D11: 弹幕量
        ws.cell(11, 5, 200)     # E11: 评论量
        ws.cell(11, 6, 1000)    # F11: 硬币量
        ws.cell(11, 7, 1500)   # G11: 收藏量
        
        ws.cell(12, 2, 50000)   # B12: 播放量
        ws.cell(12, 3, 3000)    # C12: 点赞量
        ws.cell(12, 4, 600)     # D12: 弹幕量
        ws.cell(12, 5, 300)     # E12: 评论量
        ws.cell(12, 6, 1000)    # F12: 硬币量
        ws.cell(12, 7, 1500)   # G12: 收藏量
        
        wb.save(self.test_file)
        wb.close()
    
    def test_initialization(self):
        """测试Excel管理器初始化"""
        manager = ExcelManager(self.test_file, "Sheet1")
        
        assert manager.file_path == self.test_file
        assert manager.sheet_name == "Sheet1"
        assert manager._workbook is None
        assert manager._worksheet is None
    
    def test_open_existing_file(self):
        """测试打开现有文件"""
        manager = ExcelManager(self.test_file, "Sheet1")
        manager.open()
        
        assert manager._workbook is not None
        assert manager._worksheet is not None
        
        manager.close()
    
    def test_open_nonexistent_file(self):
        """测试打开不存在的文件"""
        nonexistent_file = Path(self.temp_dir) / "nonexistent.xlsx"
        manager = ExcelManager(nonexistent_file, "Sheet1")
        
        with pytest.raises(ExcelReadError) as exc_info:
            manager.open()
        
        assert "文件不存在" in str(exc_info.value)
    
    def test_open_invalid_sheet(self):
        """测试打开不存在的工作表"""
        manager = ExcelManager(self.test_file, "NonexistentSheet")
        
        with pytest.raises(ExcelReadError) as exc_info:
            manager.open()
        
        assert "工作表不存在" in str(exc_info.value)
    
    def test_context_manager(self):
        """测试上下文管理器"""
        with ExcelManager(self.test_file, "Sheet1") as manager:
            assert manager._workbook is not None
            assert manager._worksheet is not None
        
        # 验证文件已关闭
        assert manager._workbook is None
        assert manager._worksheet is None
    
    def test_read_base_stats(self):
        """测试读取基数数据（遍历累加）"""
        with ExcelManager(self.test_file, "Sheet1") as manager:
            base_stats = manager.read_base_stats(start_row=11)
        
        assert isinstance(base_stats, VideoStats)
        # 50000 + 50000 = 100000
        assert base_stats.view == 100000
        # 2000 + 3000 = 5000
        assert base_stats.like == 5000
        # 400 + 600 = 1000
        assert base_stats.danmaku == 1000
        # 200 + 300 = 500
        assert base_stats.reply == 500
        # 1000 + 1000 = 2000
        assert base_stats.coin == 2000
        # 1500 + 1500 = 3000
        assert base_stats.favorite == 3000
    
    def test_read_base_stats_with_default_values(self):
        """测试读取基数数据（空文件返回0）"""
        # 创建一个没有数据的Excel文件
        empty_file = Path(self.temp_dir) / "empty.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        wb.save(empty_file)
        wb.close()
        
        with ExcelManager(empty_file, "Sheet1") as manager:
            base_stats = manager.read_base_stats(start_row=11)
        
        assert isinstance(base_stats, VideoStats)
        assert base_stats.view == 0
        assert base_stats.like == 0
        assert base_stats.danmaku == 0
        assert base_stats.reply == 0
        assert base_stats.coin == 0
        assert base_stats.favorite == 0
        
        # 清理
        os.remove(empty_file)
    
    def test_read_base_stats_without_open(self):
        """测试未打开文件时读取基数数据"""
        manager = ExcelManager(self.test_file, "Sheet1")
        
        with pytest.raises(ExcelReadError) as exc_info:
            manager.read_base_stats(start_row=11)
        
        assert "Excel文件未打开" in str(exc_info.value)
    
    def test_write_stats(self):
        """测试写入统计数据（追加到空行）"""
        stats = VideoStats(
            view=200000,
            like=10000,
            danmaku=2000,
            reply=1000,
            coin=4000,
            favorite=6000
        )
        
        with ExcelManager(self.test_file, "Sheet1") as manager:
            result = manager.write_stats(stats, "2024-01-01 12:00:00", start_row=11)
        
        assert result is True
        
        # 验证数据已写入第13行（11、12行已有数据，追加到13行）
        wb = openpyxl.load_workbook(self.test_file)
        ws = wb["Sheet1"]
        
        assert ws.cell(13, 2).value == 200000  # B13: 播放量
        assert ws.cell(13, 3).value == 10000   # C13: 点赞量
        assert ws.cell(13, 4).value == 2000    # D13: 弹幕量
        assert ws.cell(13, 5).value == 1000    # E13: 评论量
        assert ws.cell(13, 6).value == 4000    # F13: 硬币量
        assert ws.cell(13, 7).value == 6000    # G13: 收藏量
        
        # 验证时间信息
        assert "数据获取时间：2024-01-01 12:00:00" in ws.cell(1, 1).value
        
        wb.close()
    
    def test_write_stats_without_open(self):
        """测试未打开文件时写入统计数据"""
        stats = VideoStats(
            view=200000,
            like=10000,
            danmaku=2000,
            reply=1000,
            coin=4000,
            favorite=6000
        )
        
        manager = ExcelManager(self.test_file, "Sheet1")
        
        with pytest.raises(ExcelReadError) as exc_info:
            manager.write_stats(stats, "2024-01-01 12:00:00", start_row=11)
        
        assert "Excel文件未打开" in str(exc_info.value)
    
    def test_write_stats_to_empty_file(self):
        """测试写入到空文件"""
        # 创建一个没有数据的Excel文件
        simple_file = Path(self.temp_dir) / "simple.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        wb.save(simple_file)
        wb.close()
        
        stats = VideoStats(
            view=200000,
            like=10000,
            danmaku=2000,
            reply=1000,
            coin=4000,
            favorite=6000
        )
        
        with ExcelManager(simple_file, "Sheet1") as manager:
            manager.write_stats(stats, "2024-01-01 12:00:00", start_row=11)
        
        # 验证合并单元格已创建且数据写入第11行
        wb = openpyxl.load_workbook(simple_file)
        ws = wb["Sheet1"]
        
        assert "A1:G1" in [str(mr) for mr in ws.merged_cells]
        assert "数据获取时间：2024-01-01 12:00:00" in ws.cell(1, 1).value
        
        # 数据写入第11行（第一个空行）
        assert ws.cell(11, 2).value == 200000
        assert ws.cell(11, 3).value == 10000
        
        wb.close()
        
        # 清理
        os.remove(simple_file)
    
    def test_write_stats_preserves_existing_data(self):
        """测试写入时保留现有数据（追加而非覆盖）"""
        stats1 = VideoStats(
            view=100000,
            like=5000,
            danmaku=1000,
            reply=500,
            coin=2000,
            favorite=3000
        )
        
        with ExcelManager(self.test_file, "Sheet1") as manager:
            manager.write_stats(stats1, "2024-01-01 12:00:00", start_row=11)
        
        # 写入新数据
        stats2 = VideoStats(
            view=200000,
            like=10000,
            danmaku=2000,
            reply=1000,
            coin=4000,
            favorite=6000
        )
        
        with ExcelManager(self.test_file, "Sheet1") as manager:
            manager.write_stats(stats2, "2024-01-02 12:00:00", start_row=11)
        
        # 验证原有数据仍在，新数据追加到下一行
        wb = openpyxl.load_workbook(self.test_file)
        ws = wb["Sheet1"]
        
        # 原有第11、12行数据不变
        assert ws.cell(11, 2).value == 50000
        assert ws.cell(12, 2).value == 50000
        
        # 第一次写入在第13行
        assert ws.cell(13, 2).value == 100000
        # 第二次写入在第14行
        assert ws.cell(14, 2).value == 200000
        
        # 验证时间信息已更新
        assert "数据获取时间：2024-01-02 12:00:00" in ws.cell(1, 1).value
        
        wb.close()
    
    def test_write_stats_with_different_sheet(self):
        """测试写入到不同工作表"""
        # 创建一个包含多个工作表的Excel文件
        multi_sheet_file = Path(self.temp_dir) / "multi_sheet.xlsx"
        wb = openpyxl.Workbook()
        
        # 创建第一个工作表
        ws1 = wb.active
        ws1.title = "Sheet1"
        
        # 创建第二个工作表
        ws2 = wb.create_sheet("Sheet2")
        
        wb.save(multi_sheet_file)
        wb.close()
        
        stats = VideoStats(
            view=200000,
            like=10000,
            danmaku=2000,
            reply=1000,
            coin=4000,
            favorite=6000
        )
        
        with ExcelManager(multi_sheet_file, "Sheet2") as manager:
            manager.write_stats(stats, "2024-01-01 12:00:00", start_row=11)
        
        # 验证数据写入到了正确的工作表
        wb = openpyxl.load_workbook(multi_sheet_file)
        ws2 = wb["Sheet2"]
        
        assert ws2.cell(11, 2).value == 200000
        assert ws2.cell(11, 3).value == 10000
        
        wb.close()
        
        # 清理
        os.remove(multi_sheet_file)
    
    def test_read_after_write(self):
        """测试先写入再读取"""
        stats = VideoStats(
            view=200000,
            like=10000,
            danmaku=2000,
            reply=1000,
            coin=4000,
            favorite=6000
        )
        
        # 写入数据（追加到第13行，因为11、12已有数据）
        with ExcelManager(self.test_file, "Sheet1") as manager:
            manager.write_stats(stats, "2024-01-01 12:00:00", start_row=11)
        
        # 读取基数数据（从第11行累加）
        with ExcelManager(self.test_file, "Sheet1") as manager:
            base_stats = manager.read_base_stats(start_row=11)
        
        # 50000 + 50000 + 200000 = 300000
        assert base_stats.view == 300000
        # 2000 + 3000 + 10000 = 15000
        assert base_stats.like == 15000
        # 400 + 600 + 2000 = 3000
        assert base_stats.danmaku == 3000
        # 200 + 300 + 1000 = 1500
        assert base_stats.reply == 1500
        # 1000 + 1000 + 4000 = 6000
        assert base_stats.coin == 6000
        # 1500 + 1500 + 6000 = 9000
        assert base_stats.favorite == 9000
    
    def test_write_stats_with_zero_values(self):
        """测试写入零值数据"""
        stats = VideoStats(
            view=0,
            like=0,
            danmaku=0,
            reply=0,
            coin=0,
            favorite=0
        )
        
        # 创建一个空文件
        empty_file = Path(self.temp_dir) / "zero.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        wb.save(empty_file)
        wb.close()
        
        with ExcelManager(empty_file, "Sheet1") as manager:
            result = manager.write_stats(stats, "2024-01-01 12:00:00", start_row=11)
        
        assert result is True
        
        # 验证数据已写入
        wb = openpyxl.load_workbook(empty_file)
        ws = wb["Sheet1"]
        
        assert ws.cell(11, 2).value == 0
        assert ws.cell(11, 3).value == 0
        assert ws.cell(11, 4).value == 0
        assert ws.cell(11, 5).value == 0
        assert ws.cell(11, 6).value == 0
        assert ws.cell(11, 7).value == 0
        
        wb.close()
        
        # 清理
        os.remove(empty_file)
    
    def test_write_stats_with_large_values(self):
        """测试写入大数值数据"""
        stats = VideoStats(
            view=10000000,
            like=500000,
            danmaku=100000,
            reply=50000,
            coin=200000,
            favorite=300000
        )
        
        with ExcelManager(self.test_file, "Sheet1") as manager:
            result = manager.write_stats(stats, "2024-01-01 12:00:00", start_row=11)
        
        assert result is True
        
        # 验证数据已写入第13行
        wb = openpyxl.load_workbook(self.test_file)
        ws = wb["Sheet1"]
        
        assert ws.cell(13, 2).value == 10000000
        assert ws.cell(13, 3).value == 500000
        assert ws.cell(13, 4).value == 100000
        assert ws.cell(13, 5).value == 50000
        assert ws.cell(13, 6).value == 200000
        assert ws.cell(13, 7).value == 300000
        
        wb.close()


class TestExcelErrorClasses:
    """Excel异常类测试"""
    
    def test_excel_error_hierarchy(self):
        """测试Excel异常类层次结构"""
        assert issubclass(ExcelReadError, ExcelError)
        assert issubclass(ExcelWriteError, ExcelError)
        assert issubclass(ExcelError, Exception)
    
    def test_excel_read_error(self):
        """测试Excel读取异常"""
        error = ExcelReadError("读取失败")
        assert str(error) == "读取失败"
        assert isinstance(error, ExcelError)
    
    def test_excel_write_error(self):
        """测试Excel读取异常"""
        error = ExcelWriteError("写入失败")
        assert str(error) == "写入失败"
        assert isinstance(error, ExcelError)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
