import tkinter as tk
from tkinter import messagebox
import requests
import math
from datetime import datetime, timezone, timedelta
import email.utils
import openpyxl

# Windows 字体 DPI 优化
try:
    from ctypes import windll
    windll.shcore.SetProcessDpiAwareness(1)
except:
    pass

# ====================== Excel 操作 ======================
def read_base_from_excel(file_path, sheet_name):
    """读取 B5:G5 的累计基数，返回字典"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        base = {
            'view': ws.cell(5, 2).value or 0,
            'like': ws.cell(5, 3).value or 0,
            'danmaku': ws.cell(5, 4).value or 0,
            'reply': ws.cell(5, 5).value or 0,
            'coin': ws.cell(5, 6).value or 0,
            'favorite': ws.cell(5, 7).value or 0
        }
        wb.close()
        return base
    except Exception as e:
        messagebox.showerror("读取基数失败", f"无法读取 Excel 中的基数：{e}")
        return None

def write_raw_to_excel(file_path, sheet_name, stats, beijing_time_str):
    """将总原始数据固定写入 B4:G4，并将北京时间写入 A1:G1 合并单元格"""
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]

        # --- 新增：处理 A1:G1 合并单元格，写入北京时间 ---
        # 检查是否已合并，若未合并则执行合并
        if 'A1:G1' not in [str(mr) for mr in ws.merged_cells]:
            ws.merge_cells('A1:G1')
        ws.cell(1, 1).value = f"数据获取时间：{beijing_time_str}"

        # --- 原有写入 B4:G4 的代码（保持不变）---
        row = 4
        merged = list(ws.merged_cells)
        def is_merged(r, c):
            for mr in merged:
                if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
                    return True
            return False
        for col in range(2, 8):
            if is_merged(row, col):
                messagebox.showerror("写入失败", f"第{row}行第{col}列属于合并单元格")
                return False
        ws.cell(row, 2, stats['view'])
        ws.cell(row, 3, stats['like'])
        ws.cell(row, 4, stats['danmaku'])
        ws.cell(row, 5, stats['reply'])
        ws.cell(row, 6, stats['coin'])
        ws.cell(row, 7, stats['favorite'])
        wb.save(file_path)
        return True
    except Exception as e:
        messagebox.showerror("写入失败", f"写入 Excel 出错：{e}")
        return False

# ====================== API 获取 ======================
def get_video_stats(bvid):
    url = f"https://api.bilibili.com/x/web-interface/view?bvid={bvid}"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
    try:
        resp = requests.get(url, headers=headers, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        if data['code'] != 0:
            raise Exception(f"API返回错误: {data['message']}")
        stat = data['data']['stat']
        server_time = resp.headers.get('Date')
        return {
            'view': stat['view'],
            'like': stat['like'],
            'danmaku': stat['danmaku'],
            'reply': stat['reply'],
            'coin': stat['coin'],
            'favorite': stat['favorite'],
            'server_time': server_time
        }
    except Exception as e:
        messagebox.showerror("获取失败", str(e))
        return None

# ====================== 得分计算（基于增量） ======================
def compute_weekly_score(stats):
    """stats 为增量数据字典"""
    A3 = float(stats['view'])
    B3 = float(stats['like'])
    C3 = float(stats['danmaku'])
    D3 = float(stats['reply'])
    E3 = float(stats['coin'])
    F3 = float(stats['favorite'])

    A6 = A3 * 0.5 + 5000 if A3 > 10000 else A3
    B6 = D3 + C3
    denominator = A6 + F3 + B6 * 20
    C6 = 0 if denominator == 0 else ((A6 + F3) / denominator) ** 2

    # 修正B
    if F3 > E3 * 2:
        temp = (E3 ** 2 / (A3 * F3)) * 1000 if A3 * F3 != 0 else 0
    else:
        temp = (F3 / A3) * 250 if A3 != 0 else 0
    D6 = 50 if temp > 50 else temp

    # 修正C
    if E3 > F3:
        temp = (F3 ** 2 / (A3 * E3)) * 250 if A3 * E3 != 0 else 0
    else:
        temp = (E3 / A3) * 250 if A3 != 0 else 0
    E6 = 50 if temp > 50 else temp

    # 修正D
    if F3 > E3:
        temp = (E3 / A3) * 25 if A3 != 0 else 0
    else:
        temp = (F3 / A3) * 25 if A3 != 0 else 0
    F6 = 1 if temp > 1 else temp

    play = A6 * F6
    interaction = B6 * C6 * 15
    favorite_pt = F3 * D6
    coin_pt = E3 * E6
    like_pt = min(B3, E3 * 2)
    total = play + interaction + favorite_pt + coin_pt + like_pt

    return {
        "播放得点": play,
        "互动得点": interaction,
        "收藏得点": favorite_pt,
        "硬币得点": coin_pt,
        "点赞得点": like_pt,
        "最终得点": total
    }

# ====================== GUI 三列布局 ======================
class WeeklyScoreApp:
    def __init__(self, root):
        self.root = root
        self.root.title("B站周刊得点计算器")
        self.root.geometry("850x400")
        self.root.minsize(850, 400)

        self.sheet_name = "Sheet1"

        font_label = ("楷体", 10)
        font_data = ("微软雅黑", 9)

        # ----- 顶部 -----
        top_frame = tk.Frame(root)
        top_frame.pack(pady=10, fill=tk.X)
        left_part = tk.Frame(top_frame)
        left_part.pack(side=tk.LEFT)
        tk.Label(left_part, text=" BV号：", font=font_label).pack(side=tk.LEFT, padx=5)
        self.bvid_entry = tk.Entry(left_part, width=18, font=("等线", 10))
        self.bvid_entry.pack(side=tk.LEFT, padx=5)
        self.query_btn = tk.Button(left_part, text="查询", font=("楷体", 10), width=8, command=self.query_data)
        self.query_btn.pack(side=tk.LEFT, padx=10)
        self.time_label = tk.Label(top_frame, text="北京时间：--", font=("仿宋", 9))
        self.time_label.pack(side=tk.LEFT, padx=15)

        # ----- 主体：三列（使用 grid 布局）-----
        main_frame = tk.Frame(root)
        main_frame.pack(pady=10, fill=tk.BOTH, expand=True)
        # 三列等宽
        for col in range(3):
            main_frame.columnconfigure(col, weight=1)

        # 列1：总原始数据
        col1 = tk.LabelFrame(main_frame, text="▍ 总原始数据", font=("黑体", 11, "bold"), padx=10, pady=8)
        col1.grid(row=0, column=0, padx=8, sticky="nsew")
        # 列2：增量数据
        col2 = tk.LabelFrame(main_frame, text="▍ 本期增量数据", font=("黑体", 11, "bold"), padx=10, pady=8)
        col2.grid(row=0, column=1, padx=8, sticky="nsew")
        # 列3：周刊得点（增量计算结果）
        col3 = tk.LabelFrame(main_frame, text="▍ 本期周刊得点", font=("黑体", 11, "bold"), padx=10, pady=8)
        col3.grid(row=0, column=2, padx=8, sticky="nsew")

        # 存储各列的变量字典
        self.total_vars = {}
        self.delta_vars = {}
        self.score_vars = {}

        # 字段列表（顺序一致）
        fields = [
            ("播放量", "view"), ("点赞量", "like"), ("弹幕量", "danmaku"),
            ("评论量", "reply"), ("硬币量", "coin"), ("收藏量", "favorite")
        ]
        for i, (ch, en) in enumerate(fields):
            # 列1
            tk.Label(col1, text=ch, font=font_label, anchor="e", width=8).grid(row=i, column=0, padx=5, pady=5, sticky="e")
            var_total = tk.StringVar(value="--")
            tk.Label(col1, textvariable=var_total, font=font_data, width=12, anchor="w", relief="sunken", bd=1, bg="white").grid(row=i, column=1, padx=5, pady=5)
            self.total_vars[en] = var_total

            # 列2
            tk.Label(col2, text=ch, font=font_label, anchor="e", width=8).grid(row=i, column=0, padx=5, pady=5, sticky="e")
            var_delta = tk.StringVar(value="--")
            tk.Label(col2, textvariable=var_delta, font=font_data, width=12, anchor="w", relief="sunken", bd=1, bg="white").grid(row=i, column=1, padx=5, pady=5)
            self.delta_vars[en] = var_delta

        # 列3：得分项
        score_fields = [
            ("播放得点", "play"), ("互动得点", "interaction"),
            ("收藏得点", "favorite_points"), ("硬币得点", "coin_points"),
            ("点赞得点", "like_points"), ("最终得点", "total")
        ]
        for i, (ch, en) in enumerate(score_fields):
            tk.Label(col3, text=ch, font=font_label, anchor="e", width=8).grid(row=i, column=0, padx=5, pady=5, sticky="e")
            var_score = tk.StringVar(value="--")
            tk.Label(col3, textvariable=var_score, font=font_data, width=12, anchor="w", relief="sunken", bd=1, bg="white").grid(row=i, column=1, padx=5, pady=5)
            self.score_vars[en] = var_score

        # 底部说明
        bottom_frame = tk.Frame(root)
        bottom_frame.pack(pady=8)
        tk.Label(bottom_frame, text="请将计算器与Excel文件放在同一文件夹内，在表格中准确填写历史数据", font=("仿宋", 8), fg="gray").pack()
        tk.Label(bottom_frame, text="相关计算公式来源于B站中V周刊组，本计算器由云师阳(1866643210)制作", font=("仿宋", 8), fg="gray").pack()

    def query_data(self):
        bvid = self.bvid_entry.get().strip()
        if not bvid:
            messagebox.showwarning("警告", "请输入 BV 号")
            return

        # 读取基数
        excel_file = f"{bvid}.xlsx"
        base = read_base_from_excel(excel_file, self.sheet_name)
        if base is None:
            return

        self.query_btn.config(state=tk.DISABLED)
        self.root.update()

        # 获取 API 数据
        total = get_video_stats(bvid)
        if total is None:
            self.query_btn.config(state=tk.NORMAL)
            return

        # 显示北京时间
        if total.get('server_time'):
            try:
                gmt = email.utils.parsedate_to_datetime(total['server_time'])
                bj = gmt.astimezone(timezone(timedelta(hours=8)))
                self.time_label.config(text=f"北京时间：{bj.strftime('%Y-%m-%d %H:%M:%S')}")
            except:
                self.time_label.config(text="北京时间：解析失败")
        else:
            self.time_label.config(text="北京时间：未获取")

        # 更新总原始数据列（显示累计）
        self.total_vars['view'].set(f"{total['view']:,}")
        self.total_vars['like'].set(f"{total['like']:,}")
        self.total_vars['danmaku'].set(f"{total['danmaku']:,}")
        self.total_vars['reply'].set(f"{total['reply']:,}")
        self.total_vars['coin'].set(f"{total['coin']:,}")
        self.total_vars['favorite'].set(f"{total['favorite']:,}")

        # 计算增量
        delta = {
            'view': max(0, total['view'] - base['view']),
            'like': max(0, total['like'] - base['like']),
            'danmaku': max(0, total['danmaku'] - base['danmaku']),
            'reply': max(0, total['reply'] - base['reply']),
            'coin': max(0, total['coin'] - base['coin']),
            'favorite': max(0, total['favorite'] - base['favorite'])
        }

        # 更新增量列
        self.delta_vars['view'].set(f"{delta['view']:,}")
        self.delta_vars['like'].set(f"{delta['like']:,}")
        self.delta_vars['danmaku'].set(f"{delta['danmaku']:,}")
        self.delta_vars['reply'].set(f"{delta['reply']:,}")
        self.delta_vars['coin'].set(f"{delta['coin']:,}")
        self.delta_vars['favorite'].set(f"{delta['favorite']:,}")

        # 计算得分（基于增量）
        if all(v == 0 for v in delta.values()):
            scores = {k: 0.0 for k in ["播放得点","互动得点","收藏得点","硬币得点","点赞得点","最终得点"]}
        else:
            scores = compute_weekly_score(delta)

        # 更新得分列
        self.score_vars['play'].set(f"{scores['播放得点']:.2f}")
        self.score_vars['interaction'].set(f"{scores['互动得点']:.2f}")
        self.score_vars['favorite_points'].set(f"{scores['收藏得点']:.2f}")
        self.score_vars['coin_points'].set(f"{scores['硬币得点']:.2f}")
        self.score_vars['like_points'].set(f"{scores['点赞得点']:.2f}")
        self.score_vars['total'].set(f"{scores['最终得点']:.2f}")

        # 将总原始数据写入 Excel B4:G4
        # 在获取北京时间后（已有 bj 变量），格式化为字符串
        beijing_time_str = bj.strftime('%Y-%m-%d %H:%M:%S') if 'bj' in locals() else "获取失败"
        write_raw_to_excel(excel_file, self.sheet_name, total, beijing_time_str)

        self.query_btn.config(state=tk.NORMAL)

if __name__ == "__main__":
    root = tk.Tk()
    app = WeeklyScoreApp(root)
    root.mainloop()
